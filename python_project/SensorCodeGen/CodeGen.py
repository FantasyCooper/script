import this

from openpyxl import load_workbook

ClassNameFile = 'ClassName.txt'
PackageNameFile = 'PackageName.txt'
PermissionFile = 'Permission.txt'
MethodReturnFile = 'ActiveMethodCall.txt'
CallBackReturnFile = 'CallBack.txt'
WorkbookFile = 'MY19-Requested-GIS-685 Vehicle Signals Specification v2.1 26JUN2015-ToBeCollectedMarking-10-06-2017-Final.xlsx'
PackageNameKey = u'SDK Signal Fully Qualified Class'
ClassNameKey = u'Class'
PermissionKey = u'Specific Read Permission'
ParaNameKey = u'Global A Short Name'
PriorityKey = u'MY19 Priority'
ReturnValueKey = u'Return Value'
ShortNameKey = u'Global A Short Name'
FrequencyCountKey = u'Frequency Count'

ListenerTemplate = "steeringWheelAngle.setOnDataChangeListener(new OnVehicleDataListener() {\n    @SuppressLint(\"LongLogTag\")\n    @Override\n    public void onVehicleDataChange(int i) {\n        try {\n            Log.d(\"steeringWheelAngle\", \"StrWhAng\" + steeringWheelAngle.getValue().StrWhAng + \", value from listener update.\");\n        } catch (VehicleDataUnavailable vehicleDataUnavailable) {\n            Log.d(\"steeringWheelAngle\", \"No valid data but listener triggered.\");\n            vehicleDataUnavailable.printStackTrace();\n        }\n    }\n});\n"


class JavaCode:
    className = set()
    packageName = set()
    permission = set()
    workBook = None
    workSheet = None
    key = dict()
    package_name_col = -1
    class_name_col = -1
    permission_col = -1
    para_name_col = -1
    priority_col = -1
    return_value_col = -1
    short_name_col = -1
    frequency_count_col = -1

    def __init__(self):
        self.workBook = load_workbook(WorkbookFile)
        self.workSheet = self.workBook.active
        rowOne = self.workSheet[1]
        # with open(ClassNameFile, 'r') as f:
        #     for line in f:
        #         if line.strip() != '':
        #             self.className.add(line.strip())
        # with open(PackageNameFile, 'r') as f:
        #     for line in f:
        #         if line.strip() != '':
        #             self.packageName.add(line.strip())
        # with open(PermissionFile, 'r') as f:
        #     for line in f:
        #         self.permission.add(line.strip())
        for i in range(0, len(rowOne)):
            self.key[rowOne[i].value] = i
        self.package_name_col = self.key[PackageNameKey]
        self.class_name_col = self.key[ClassNameKey]
        self.permission_col = self.key[PermissionKey]
        self.para_name_col = self.key[ParaNameKey]
        self.priority_col = self.key[PriorityKey]
        self.return_value_col = self.key[ReturnValueKey]
        self.short_name_col = self.key[ShortNameKey]
        self.frequency_count_col = self.key[FrequencyCountKey]

    def __del__(self):
        self.workBook.save("copy.xlsx")
    def gen_java(self):  # todo:
        permission_set = set()
        package_name_low = set()
        package_name_high = set()
        callback_java_high = ''
        callback_java_low = ''
        for row in self.workSheet.iter_rows(min_row=2):
            if row[self.priority_col].value.strip() == 'P0':
                package_name = row[self.package_name_col].value
                permission = row[self.permission_col].value
                class_name = row[self.class_name_col].value
                para_name = row[self.para_name_col].value
                if permission.strip() != '' and permission.find('<any>') == -1 and permission.find('N/A') == -1:
                    permission_set.add(permission)

    def update_active_call(self):
        method_return = dict()
        call_back_return = dict()
        with open(MethodReturnFile, 'r') as f:
            for line in f:
                full_class_name = ''
                if line.find('high CAN') != -1:
                    full_class_name += 'HIGH'
                elif line.find('low CAN') != -1:
                    full_class_name += 'LOW'
                key_pair = line[line.find(' D ') + 3: line.find(', value from')]
                full_class_name += key_pair.split(':')[0].strip()
                if line.find('active method call V') != -1:
                    full_class_name += 'V'
                method_return[full_class_name] = key_pair.split(':')[1].strip()
        with open(CallBackReturnFile, 'r') as f:
            for line in f:
                full_class_name = ''
                if line.find('high CAN') != -1:
                    full_class_name += 'HIGH'
                elif line.find('low CAN') != -1:
                    full_class_name += 'LOW'
                key_pair = line[line.find(' D ') + 3: line.find(', value from')]
                full_class_name += key_pair.split(':')[0].strip()
                if line.find('listener update V') != -1:
                    full_class_name += 'V'
                call_back_return[full_class_name] = call_back_return.get(full_class_name, 0) + 1
        # for row in self.workSheet.iter_rows(min_row=2):
        #     if row[self.priority_col].value.strip() == 'P0':
        #         package_name = row[self.package_name_col].value
        for row in self.workSheet.iter_rows(min_row=2):
            if row[self.priority_col].value == 'P0':
                class_name = row[self.class_name_col].value.encode('ascii')
                class_name = class_name[:1].lower() + class_name[1:]
                if row[self.short_name_col].value[-1] == 'V':
                    class_name += 'V'
                if row[self.package_name_col].value.find('hsgmlan') != -1:
                    class_name = 'HIGH' + class_name
                elif row[self.package_name_col].value.find('lsgmlan') != -1:
                    class_name = 'LOW' + class_name
                value = method_return.get(class_name, "No return value")
                count = call_back_return.get(class_name, 0)
                row[self.return_value_col].value = value
                row[self.frequency_count_col].value = count

    def gen_permission(self):
        permission_java = ''
        for item in self.permission:
            temp = '    <uses-permission android:name=\"' + item + '\"/>\n'
            permission_java += temp
        return permission_java

    def gen_import(self):
        import_java = ''
        for item in self.packageName:
            temp = 'import ' + item + ';\n'
            import_java += temp
        return import_java

    def gen_callback(self):
        var_def = ''
        var_init = ''
        set_listener = ''
        for item in self.className:
            de_capital = item[:1].lower() + item[1:]
            # Generate defination. eg private SteeringWheelAngle steeringWheelAngle;
            var_def += 'private ' + item + ' ' + de_capital + ';\n'
            # Generate initialize statement. eg steeringWheelAngle = new SteeringWheelAngle(context);
            var_init += de_capital + ' = new ' + item + '(context);\n'
            # Generate set listener statement, template below is Java statement for steeringWheelAngle
            template = 'steeringWheelAngle.setOnDataChangeListener(new OnVehicleDataListener() {\r\n    @SuppressLint(\"LongLogTag\")\r\n    @Override\r\n    public void onVehicleDataChange(int i) {\r\n        try {\r\n            Log.d(\"steeringWheelAngle\", steeringWheelAngle.getValue() + \", value from listener update.\");\r\n        } catch (VehicleDataUnavailable vehicleDataUnavailable) {\r\n            Log.d(\"steeringWheelAngle\", \"No valid data but listener triggered.\");\r\n            vehicleDataUnavailable.printStackTrace();\r\n        }\r\n    }\r\n});\r\n'
            set_listener += template.replace('steeringWheelAngle', de_capital)
        return var_def + '\n' + var_init + '\n' + set_listener


javaCode = JavaCode()
javaCode.update_active_call()
# print javaCode.gen_permission()
# print '\n\n\n'
# print javaCode.gen_import()
# print '\n\n\n'
# print javaCode.gen_callback()
